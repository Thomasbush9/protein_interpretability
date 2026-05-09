#!/usr/bin/env python3
"""Build Boltz YAML inputs for the Porter et al. (2024) fold-switching benchmark.

Reads TableS1 from the ncbi/AF2_benchmark repo, fetches full chain sequences
from the RCSB PDB FASTA API, and writes one Boltz YAML per fold-switcher pair
(~99 total) + a manifest.csv.

Reference: Chakravarty, Schafer, Chen, Cotten, Porter (2024).
"AlphaFold predictions of fold-switched conformations are driven by structure
memorization." Nat Commun 15:7296. https://doi.org/10.1038/s41467-024-51801-z

The paper's TableS1 lists Fold1/Fold2 PDB+chain pairs (e.g. "1h38D" = PDB 1H38
chain D) plus the *fold-switching region* sequence (~20-40 residues). Boltz
needs the *full* chain sequence, which we fetch from RCSB.

Usage::

    python scripts/foldswitch_setup.py \\
        --table_s1 "/tmp/AF2_benchmark/supporting tables/TableS1.xlsx" \\
        --out_dir  /path/to/foldswitch \\
        [--msa_mode auto|empty]    # auto = omit msa: field (Boltz auto-MSA)
                                    # empty = msa: empty (single-sequence ablation)

Output::

    <out_dir>/
        manifest.csv             # one row per pair (idx, fold1, fold2, chains,
                                 #                   seq_len, seq_status, fs_region)
        yamls/
            seq_0001_<fold1>_<fold2>.yaml
            ...
        cache/                   # cached RCSB FASTA responses (re-runs are free)
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import urlopen

import openpyxl

RCSB_FASTA_URL = "https://www.rcsb.org/fasta/entry/{pdb}"
PAIR_RE = re.compile(r"^([0-9a-zA-Z]{4})([A-Za-z0-9])$")


def parse_pdb_chain(token: str) -> tuple[str, str]:
    """E.g. '1h38D' -> ('1H38', 'D').  Some entries are 5+1 chars; flexible."""
    token = token.strip()
    m = PAIR_RE.match(token)
    if m:
        return m.group(1).upper(), m.group(2).upper()
    if "_" in token:
        a, b = token.rsplit("_", 1)
        return a.upper(), b.upper()
    raise ValueError(f"cannot parse PDB+chain from {token!r}")


def read_table_s1(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    pairs = []
    for i, r in enumerate(rows[1:], start=1):
        f1, f2, fs_region = r[0], r[1], r[2]
        if not f1 or not f2:
            continue  # trailing blanks
        pdb1, ch1 = parse_pdb_chain(str(f1))
        pdb2, ch2 = parse_pdb_chain(str(f2))
        pairs.append(dict(
            idx=i, fold1=str(f1), fold2=str(f2),
            pdb1=pdb1, ch1=ch1, pdb2=pdb2, ch2=ch2,
            fs_region=str(fs_region) if fs_region else "",
        ))
    return pairs


def fetch_fasta(pdb: str, cache_dir: Path) -> str:
    cache_file = cache_dir / f"{pdb}.fasta"
    if cache_file.exists():
        return cache_file.read_text()
    url = RCSB_FASTA_URL.format(pdb=pdb)
    for attempt in range(3):
        try:
            with urlopen(url, timeout=20) as resp:
                txt = resp.read().decode("utf-8")
            cache_file.write_text(txt)
            return txt
        except (HTTPError, URLError, TimeoutError) as e:
            if attempt == 2:
                raise RuntimeError(f"RCSB fetch failed for {pdb}: {e}") from e
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"unreachable for {pdb}")


def parse_fasta_records(txt: str) -> list[tuple[str, str]]:
    """Returns list of (header, sequence)."""
    records, header, body = [], None, []
    for line in txt.splitlines():
        if line.startswith(">"):
            if header is not None:
                records.append((header, "".join(body)))
            header, body = line[1:], []
        else:
            body.append(line.strip())
    if header is not None:
        records.append((header, "".join(body)))
    return records


def chain_seq_from_fasta(fasta_txt: str, chain: str) -> str | None:
    """Find the sequence whose record header lists `chain` as an auth_asym_id.

    RCSB FASTA headers look like::

        >1H38_1|Chain A[auth A]|...
        >1H38_2|Chains B, C[auth B], [auth C]|...

    We match `[auth <CHAIN>]` (case-insensitive) and fall back to bare 'Chain X'
    listings if no auth annotation is present.
    """
    chain_u = chain.upper()
    for header, seq in parse_fasta_records(fasta_txt):
        # Look for explicit [auth X] annotations
        auth_ids = re.findall(r"\[auth\s+([A-Za-z0-9]+)\]", header)
        if auth_ids:
            if chain_u in {a.upper() for a in auth_ids}:
                return seq
            continue
        # Fallback: bare 'Chain X' / 'Chains X, Y' tokens (no auth remap)
        m = re.search(r"Chains?\s+([A-Za-z0-9, ]+?)\|", header)
        if m:
            chains = {c.strip().upper() for c in m.group(1).split(",")}
            if chain_u in chains:
                return seq
    return None


def resolve_pair(pair: dict, cache_dir: Path) -> dict:
    out = dict(pair)
    try:
        f1_txt = fetch_fasta(pair["pdb1"], cache_dir)
        f2_txt = fetch_fasta(pair["pdb2"], cache_dir)
        s1 = chain_seq_from_fasta(f1_txt, pair["ch1"])
        s2 = chain_seq_from_fasta(f2_txt, pair["ch2"])
    except Exception as e:
        out.update(seq_status=f"error: {e}", sequence="", seq_len=0)
        return out
    if s1 is None and s2 is None:
        out.update(seq_status="no_chain_match", sequence="", seq_len=0)
        return out
    if s1 is None:
        out.update(seq_status="fold1_missing_use_fold2", sequence=s2, seq_len=len(s2))
        return out
    if s2 is None:
        out.update(seq_status="fold2_missing_use_fold1", sequence=s1, seq_len=len(s1))
        return out
    if s1 == s2:
        out.update(seq_status="identical", sequence=s1, seq_len=len(s1))
    elif s1 in s2 or s2 in s1:
        # one is a substring of the other: take longer (full SEQRES)
        longer = s1 if len(s1) >= len(s2) else s2
        out.update(seq_status="substring_use_longer",
                   sequence=longer, seq_len=len(longer))
    else:
        # Sequences differ in interior — take fold1 as canonical, flag.
        out.update(seq_status="diverge_use_fold1",
                   sequence=s1, seq_len=len(s1),
                   alt_sequence=s2, alt_seq_len=len(s2))
    return out


def write_yaml(out_path: Path, sequence: str, msa_mode: str) -> None:
    lines = [
        "version: 1",
        "sequences:",
        "  - protein:",
        "      id: A",
        f"      sequence: {sequence}",
    ]
    if msa_mode == "empty":
        lines.append("      msa: empty")
    elif msa_mode == "auto":
        pass  # omit field — Boltz auto-generates
    else:
        raise ValueError(f"msa_mode must be auto|empty, got {msa_mode!r}")
    out_path.write_text("\n".join(lines) + "\n")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--table_s1", type=Path, required=True,
                    help="Path to TableS1.xlsx from ncbi/AF2_benchmark")
    ap.add_argument("--out_dir", type=Path, required=True)
    ap.add_argument("--msa_mode", choices=["auto", "empty"], default="auto",
                    help="auto (omit msa: → Boltz generates) or empty (single-seq)")
    ap.add_argument("--max_workers", type=int, default=8)
    ap.add_argument("--limit", type=int, default=None,
                    help="Process only first N pairs (debugging)")
    args = ap.parse_args()

    out_dir: Path = args.out_dir
    yaml_dir = out_dir / "yamls"
    cache_dir = out_dir / "cache"
    yaml_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    pairs = read_table_s1(args.table_s1)
    if args.limit:
        pairs = pairs[: args.limit]
    print(f"[info] {len(pairs)} fold-switcher pairs in TableS1")

    resolved: list[dict] = []
    with ThreadPoolExecutor(max_workers=args.max_workers) as pool:
        futs = {pool.submit(resolve_pair, p, cache_dir): p for p in pairs}
        for fut in as_completed(futs):
            r = fut.result()
            resolved.append(r)
            tag = r["seq_status"]
            print(f"  [{r['idx']:>3}] {r['fold1']} / {r['fold2']:<8} "
                  f"len={r['seq_len']:>4}  {tag}")

    resolved.sort(key=lambda r: r["idx"])

    # Write YAMLs for everything that has a sequence
    n_yaml = 0
    for r in resolved:
        if not r.get("sequence"):
            continue
        stem = f"seq_{r['idx']:04d}_{r['fold1']}_{r['fold2']}"
        write_yaml(yaml_dir / f"{stem}.yaml", r["sequence"], args.msa_mode)
        r["yaml_path"] = str(yaml_dir / f"{stem}.yaml")
        n_yaml += 1

    # Manifest CSV
    csv_path = out_dir / "manifest.csv"
    fieldnames = ["idx", "fold1", "fold2", "pdb1", "ch1", "pdb2", "ch2",
                  "seq_status", "seq_len", "alt_seq_len",
                  "fs_region", "sequence", "yaml_path"]
    with csv_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in resolved:
            w.writerow({k: r.get(k, "") for k in fieldnames})

    # JSON sidecar with full info (incl. alt sequences for diverge cases)
    (out_dir / "manifest.json").write_text(json.dumps(resolved, indent=2))

    # Summary
    by_status = {}
    for r in resolved:
        by_status[r["seq_status"]] = by_status.get(r["seq_status"], 0) + 1
    print("\n[summary]")
    print(f"  pairs processed:   {len(resolved)}")
    print(f"  yamls written:     {n_yaml}  -> {yaml_dir}")
    print(f"  manifest:          {csv_path}")
    print(f"  by seq_status:")
    for k, v in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"    {k:>30}  {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
